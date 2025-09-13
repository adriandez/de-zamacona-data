#!/usr/bin/env python3
# Fase 1: preparar dataset para normalización
# - Crear columnas __work (sin duplicar __raw)
# - Quitar columnas ruido
# - Añadir link a FamilySearch por arkId (hipervínculo clickable)
# - Añadir columna childSpouseFullName (vacía, lista para rellenar)
# Entrada:  out/Zamacona_all_raw.xlsx
# Salida:   out/Zamacona_prepared.xlsx  (y CSV)

import re
import unicodedata
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

IN_FILE  = "out/Zamacona_all_raw.xlsx"
OUT_DIR  = Path("out")
OUT_XLSX = OUT_DIR / "Zamacona_prepared.xlsx"
OUT_CSV  = OUT_DIR / "Zamacona_prepared.csv"

# columnas para trabajar (single y multi)
SINGLE_COLS = ["fullName", "fatherFullName", "motherFullName", "spouseFullName"]
MULTI_COLS  = ["childrenFullNames", "otherFullNames"]  # valores separados por ';'

# columnas a quitar
DROP_COLS = [
    "sourceMediaType", "relationshipToHead", "residenceDate", "residencePlaceText",
    "parentFullNames", "otherEvents",
]

# activar normalización ligera sobre __work si quieres
DO_MINI_NORM = False

def mini_norm(s: str) -> str:
    if not isinstance(s, str): return ""
    s = s.replace("_", " ")
    s = re.sub(r"\s+", " ", s.strip())
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    return s

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if not Path(IN_FILE).exists():
        raise SystemExit(f"No encuentro {IN_FILE}")

    # leer como texto y limpiar nombres de columnas
    df = pd.read_excel(IN_FILE, dtype=str)
    df.columns = [re.sub(r"\s+", " ", str(c)).strip() for c in df.columns]

    # arkId válido
    if "arkId" in df.columns:
        df["arkId"] = df["arkId"].astype(str)
        df = df[df["arkId"].str.startswith("ark:", na=False)]
    else:
        df["arkId"] = pd.NA

    # SINGLE: crear solo __work (junto a cada columna original)
    for col in SINGLE_COLS:
        if col not in df.columns:
            df[col] = pd.NA
        work_col = f"{col}__work"
        df.insert(df.columns.get_loc(col) + 1, work_col,
                  df[col].map(mini_norm) if DO_MINI_NORM else df[col])

    # MULTI: crear __items, __count y __work (junto a cada original)
    for col in MULTI_COLS:
        if col not in df.columns:
            df[col] = pd.NA
        items = df[col].fillna("").astype(str).str.split(";")
        items = items.apply(lambda L: [re.sub(r"\s+", " ", x.strip()) for x in L if x and x.strip()] if isinstance(L, list) else [])
        items_str = items.apply(lambda L: "; ".join(L))
        count = items.apply(len)
        work = items_str.map(mini_norm) if DO_MINI_NORM else items_str
        pos = df.columns.get_loc(col) + 1
        df.insert(pos,     f"{col}__items", items_str)
        df.insert(pos + 1, f"{col}__count", count)
        df.insert(pos + 2, f"{col}__work",  work)

    # URL + quitar columnas ruido
    df["arkUrl"] = df["arkId"].apply(
        lambda x: f"https://www.familysearch.org/{x}?lang=es" if isinstance(x, str) and x.startswith("ark:") else ""
    )
    df = df.drop(columns=[c for c in DROP_COLS if c in df.columns])

    # nueva columna vacía para rellenar luego
    df["childSpouseFullName"] = ""

    # guardar base
    df.to_excel(OUT_XLSX, index=False)
    df.to_csv(OUT_CSV, index=False)
    print(f"OK -> {OUT_XLSX}  ({len(df)} filas)")
    print(f"OK -> {OUT_CSV}")

    # hipervínculos en arkUrl (pero sin crear arkLink extra)
    wb = load_workbook(OUT_XLSX)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    col_url = headers.get("arkUrl")
    if col_url:
        for r in range(2, ws.max_row + 1):
            url = (ws.cell(row=r, column=col_url).value or "").strip()
            if url:
                cell_url = ws.cell(row=r, column=col_url)
                cell_url.hyperlink = url
                cell_url.style = "Hyperlink"
        wb.save(OUT_XLSX)
        print("Hipervínculos aplicados en arkUrl.")

if __name__ == "__main__":
    main()