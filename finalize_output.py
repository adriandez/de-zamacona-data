#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
finalize_output.py
Crea el fichero FINAL con colores e hipervínculos después del pipeline.

Entrada (por prioridad):
  1) out/Zamacona_normalized_patched.xlsx
  2) out/Zamacona_normalized.xlsx

Salida:
  - out/Zamacona_final.xlsx   (repintado + hyperlinks)
  - out/Zamacona_final.csv    (opcional, sin estilos)
"""

from __future__ import annotations
from pathlib import Path
import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
OUT  = ROOT / "out"
IN_PATCHED = OUT / "Zamacona_normalized_patched.xlsx"
IN_NORMAL  = OUT / "Zamacona_normalized.xlsx"

OUT_XLSX = OUT / "Zamacona_final.xlsx"
OUT_CSV  = OUT / "Zamacona_final.csv"  # si no lo quieres, comenta la línea correspondiente

# Colores (mismos que normalize_names.py)
FILL_GREEN  = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_GRAY   = PatternFill(fill_type="solid", start_color="E7E6E6", end_color="E7E6E6")

def pick_input() -> Path:
    if IN_PATCHED.exists():
        return IN_PATCHED
    if IN_NORMAL.exists():
        return IN_NORMAL
    print("[ERROR] No encuentro normalized en out/. Ejecuta el pipeline primero.", file=sys.stderr)
    sys.exit(1)

def add_ark_hyperlinks(ws) -> int:
    # Busca cabecera 'arkId' y crea links
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    cidx = headers.get("arkId")
    if not cidx:
        return 0
    letter = get_column_letter(cidx)
    count = 0
    for r in range(2, ws.max_row + 1):
        cell = ws[f"{letter}{r}"]
        val = str(cell.value).strip() if cell.value is not None else ""
        if val.startswith("ark:/61903/1:1:"):
            cell.hyperlink = f"https://www.familysearch.org/{val}?lang=es"
            cell.style = "Hyperlink"
            count += 1
    return count

def infer_status_from_flags(df: pd.DataFrame) -> pd.Series:
    # Crea una serie 'status' a partir de flags cuando no existe
    b = df.get("blacklistFlag")
    r = df.get("reviewFlag")
    if b is None or r is None:
        # Todo verde si no hay flags (caso extremo)
        return pd.Series(["green"] * len(df))
    b = b.astype(str).fillna("0")
    r = r.astype(str).fillna("0")
    out = []
    for bb, rr in zip(b, r):
        if bb == "1":
            out.append("gray:blacklist")
        elif rr == "1":
            out.append("yellow:review")
        else:
            out.append("green")
    return pd.Series(out)

def paint_rows(ws) -> tuple[int,int,int]:
    # Decide colores usando 'status' si existe; si no, con flags
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    c_status = headers.get("status")
    c_black  = headers.get("blacklistFlag")
    c_rev    = headers.get("reviewFlag")

    def to_str(v):
        try:
            return str(v).strip().lower()
        except Exception:
            return ""

    green = yellow = gray = 0
    for r in range(2, ws.max_row + 1):
        fill = FILL_GREEN
        if c_status:
            sval = to_str(ws.cell(row=r, column=c_status).value)
            if sval.startswith("gray"):
                fill = FILL_GRAY; gray += 1
            elif sval.startswith("yellow"):
                fill = FILL_YELLOW; yellow += 1
            else:
                fill = FILL_GREEN; green += 1
        else:
            is_black = to_str(ws.cell(row=r, column=c_black).value) == "1" if c_black else "0"
            is_rev   = to_str(ws.cell(row=r, column=c_rev).value) == "1" if c_rev else "0"
            if is_black == "1" or is_black is True:
                fill = FILL_GRAY; gray += 1
            elif is_rev == "1" or is_rev is True:
                fill = FILL_YELLOW; yellow += 1
            else:
                fill = FILL_GREEN; green += 1

        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill
    return green, yellow, gray

def main() -> int:
    src = pick_input()
    # Reescribe a XLSX para asegurar que podemos repintar con openpyxl
    df = pd.read_excel(src, dtype=str)
    df.to_excel(OUT_XLSX, index=False)
    # CSV opcional (sin estilos)
    df.to_csv(OUT_CSV, index=False)

    wb = load_workbook(OUT_XLSX, data_only=False)
    ws = wb.active
    g, y, gr = paint_rows(ws)
    links = add_ark_hyperlinks(ws)
    wb.save(OUT_XLSX)

    print(f"[OK] {OUT_XLSX.name} repintado → green={g}, yellow={y}, gray={gr} (links={links})")
    print(f"[OK] {OUT_CSV.name} (sin estilos)")
    return 0

if __name__ == "__main__":
    sys.exit(main())