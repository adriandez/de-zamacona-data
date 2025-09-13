#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
find_zamacona_in_non_green.py

Promueve a 'green' filas NO green cuando:
  - Hay match estricto de 'zamacona' como palabra completa en el nombre normalizado, o
  - Hay match de sinónimo 'strong' en data/surname_synonyms.csv,
y NO hay términos de blacklist, ni ambigüedad (p.ej. 'zamacona' y 'zamacola' a la vez).

Preserva filas ya green. Colores se aplican al final según columna 'status'.
Genera:
  - out/Zamacona_normalized_patched.xlsx
  - out/Zamacona_normalized_patched.csv
  - out/Zamacona_force_green.tsv  (solo filas realmente forzadas en esta pasada)

Requisitos: pandas, openpyxl.
"""

import os
import sys
import re
import csv
import unicodedata
from pathlib import Path

import pandas as pd

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
except Exception as e:
    print("[WARN] openpyxl no disponible para colorear. Se seguirá sin color.", file=sys.stderr)
    load_workbook = None

# ---------------------------
# Config / paths
# ---------------------------

ROOT = Path(__file__).resolve().parent
DATA_DIRS = [ROOT / "data", ROOT]  # buscamos primero en data/
OUT_DIR = ROOT / "out"
OUT_DIR.mkdir(exist_ok=True)

# ficheros de entrada base
BASE_NORMALIZED = OUT_DIR / "Zamacona_normalized.xlsx"  # generado por normalize_names.py
BASE_PATCHED_PREV = OUT_DIR / "Zamacona_normalized_patched.xlsx"  # si existe, lo usamos para heredar 'status'

# ficheros auxiliares (opcionales)
SURNAME_SYNS_CSV = None
WHITELIST_TXT = None
REJECT_TXT = None

for d in DATA_DIRS:
    if (d / "surname_synonyms.csv").exists() and SURNAME_SYNS_CSV is None:
        SURNAME_SYNS_CSV = d / "surname_synonyms.csv"
    if (d / "whitelist_surnames.txt").exists() and WHITELIST_TXT is None:
        WHITELIST_TXT = d / "whitelist_surnames.txt"
    if (d / "reject_surnames.txt").exists() and REJECT_TXT is None:
        REJECT_TXT = d / "reject_surnames.txt"

# ficheros de salida
OUT_XLSX = OUT_DIR / "Zamacona_normalized_patched.xlsx"
OUT_CSV = OUT_DIR / "Zamacona_normalized_patched.csv"
OUT_FORCE = OUT_DIR / "Zamacona_force_green.tsv"

# ---------------------------
# Utils
# ---------------------------

def norm(s: str) -> str:
    if pd.isna(s):
        return ""
    s = str(s)
    s = s.lower()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^a-z\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def load_listfile(path: Path) -> set:
    vals = set()
    if path and path.exists():
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                t = line.strip()
                if t and not t.startswith("#"):
                    vals.add(norm(t))
    return vals

def load_synonyms(path: Path) -> tuple[set, set]:
    """
    Devuelve:
      strong: set de sinónimos "fuertes"
      weak: set de sinónimos "débiles" (no se usan para forzar)

    Acepta CSV con:
      - separador auto (',', ';' o tab),
      - líneas con 1..N columnas,
      - comentarios con '#',
      - columnas con nombres variados (synonym|alias|value|surname, strength|type|class|tier|level),
      - comas sin comillar: intenta deducir "strong/weak" si aparece en alguna columna.
    """
    strong, weak = set(), set()
    if not (path and path.exists()):
        return strong, weak

    import csv as _csv
    import pandas as _pd
    from pandas.errors import ParserError as _ParserError

    def _norm(s: str) -> str:
        return norm(s)  # usa tu norm() global

    # 1) Intento principal: pandas con autodetección de separador y tolerante a líneas malas
    try:
        df = _pd.read_csv(
            path,
            engine="python",
            sep=None,                  # autodetecta ',', ';', '\t'
            dtype=str,
            keep_default_na=False,
            comment="#",
            on_bad_lines="skip",       # salta líneas irrecuperables
        )
        # Detección flexible de columnas
        col_syn = next((c for c in df.columns if c.lower() in ("synonym","alias","value","surname","name")), None)
        col_str = next((c for c in df.columns if c.lower() in ("strength","type","class","tier","level")), None)

        if col_syn is None:
            # Si no hay columna clara, asume primera columna como 'synonym'
            col_syn = df.columns[0]

        if col_str is None:
            # Sin columna de fuerza: todo weak por defecto
            for v in df[col_syn]:
                vv = _norm(v)
                if vv:
                    weak.add(vv)
        else:
            for _, row in df.iterrows():
                syn = _norm(row[col_syn])
                if not syn:
                    continue
                strength = str(row[col_str]).strip().lower()
                if strength == "strong":
                    strong.add(syn)
                elif strength == "weak":
                    weak.add(syn)
                else:
                    # si la celda no es 'strong/weak', intenta deducir mirando el resto de la fila
                    tokens = {str(x).strip().lower() for x in row.tolist() if str(x).strip()}
                    if "strong" in tokens:
                        strong.add(syn)
                    elif "weak" in tokens:
                        weak.add(syn)
                    else:
                        weak.add(syn)
        return strong, weak

    except Exception as e:
        # 2) Fallback manual con csv.reader (máxima tolerancia)
        skipped = 0
        with open(path, "r", encoding="utf-8") as f:
            # intenta varios separadores
            sample = f.read()
            f.seek(0)
            # Elige separador probable
            if "\t" in sample:
                delim = "\t"
            elif ";" in sample and sample.count(";") >= sample.count(","):
                delim = ";"
            else:
                delim = ","

            reader = _csv.reader(f, delimiter=delim)
            for raw in reader:
                if not raw:
                    continue
                # une si se colaron comas sin comillas y hay más de 2 columnas
                row = [c.strip() for c in raw if c is not None]
                # salta comentarios/blank
                if len(row) == 0 or row[0].startswith("#"):
                    continue

                # Busca un posible strength en cualquier celda
                low = [c.strip().lower() for c in row]
                is_strong = any(c == "strong" for c in low)
                is_weak = any(c == "weak" for c in low)

                syn = _norm(row[0]) if row[0] else ""
                if not syn:
                    skipped += 1
                    continue

                if is_strong:
                    strong.add(syn)
                elif is_weak:
                    weak.add(syn)
                else:
                    # si hay segunda columna y parece fuerza, úsala
                    if len(row) >= 2 and row[1].strip().lower() in ("strong","weak"):
                        if row[1].strip().lower() == "strong":
                            strong.add(syn)
                        else:
                            weak.add(syn)
                    else:
                        weak.add(syn)
        if skipped:
            print(f"[WARN] load_synonyms: {skipped} líneas saltadas por vacías/sin 'synonym'.")
        return strong, weak

# ---------------------------
# Reglas de matching
# ---------------------------

ZAM_STRICT_RE = re.compile(r"\bzamacona\b")

# Blacklist mínima por defecto (se amplía con reject_surnames.txt si existe)
DEFAULT_BLACKLIST = {
    "zamacola", "zamalloa", "samacola", "samacona", "zama", "zamora"  # incluye OCRs frecuentes; amplía si ves ruido
}

def build_blacklist() -> set:
    bl = set(DEFAULT_BLACKLIST)
    rej = load_listfile(REJECT_TXT)
    bl |= rej
    return bl

def should_force(fullname_norm: str, strong_syns: set, blacklist: set) -> tuple[bool, str, bool]:
    """
    Decide si forzar a green.
    Devuelve: (force: bool, reason: str, ambiguous: bool)

    Reglas:
      - Si aparece cualquier término de blacklist -> no forzar; si también hay zamacona -> ambiguous=True
      - Si match exacto 'zamacona' -> forzar
      - Si match en sinónimo 'strong' -> forzar
      - En otro caso -> no forzar
    """
    has_bad = any(re.search(rf"\b{re.escape(b)}\b", fullname_norm) for b in blacklist)
    has_exact = bool(ZAM_STRICT_RE.search(fullname_norm))
    has_strong_syn = any(re.search(rf"\b{re.escape(s)}\b", fullname_norm) for s in strong_syns) if strong_syns else False

    if has_bad and has_exact:
        return (False, "skip:ambiguous:blacklist+exact", True)
    if has_bad and has_strong_syn:
        return (False, "skip:ambiguous:blacklist+syn", True)
    if has_bad:
        return (False, "skip:blacklist", False)

    if has_exact:
        return (True, "force:zamacona:exact", False)
    if has_strong_syn:
        return (True, "force:zamacona:syn-strong", False)

    return (False, "skip:no-match", False)

# ---------------------------
# Column helpers
# ---------------------------

def pick_first_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in candidates:
        if c in df.columns:
            return c
    # busca case-insensitive
    lc = {c.lower(): c for c in df.columns}
    for c in candidates:
        if c.lower() in lc:
            return lc[c.lower()]
    return None

def build_fullname_norm(row, cols_priority: list[str]) -> str:
    # try prioritized single column (e.g., 'fullName__work')
    for c in cols_priority:
        if c in row and pd.notna(row[c]) and str(row[c]).strip():
            return norm(str(row[c]))
    # else, try to stitch from common parts
    parts = []
    for c in ["given", "given__work", "firstName", "name"]:
        if c in row and pd.notna(row[c]):
            parts.append(str(row[c]))
    for c in ["surname", "surname1", "surname2", "lastName", "lastName1", "lastName2"]:
        if c in row and pd.notna(row[c]):
            parts.append(str(row[c]))
    if parts:
        return norm(" ".join(parts))
    # fallback to any text-ish cols
    for c in row.index:
        v = row[c]
        if isinstance(v, str) and len(v) > 0:
            return norm(v)
    return ""

# ---------------------------
# Color helpers (visual only)
# ---------------------------

FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid") if load_workbook else None
FILL_YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") if load_workbook else None
FILL_GRAY = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid") if load_workbook else None
FILL_RED = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid") if load_workbook else None

def apply_row_fill(ws, row_idx: int, status: str):
    if not load_workbook:
        return
    # colorea toda la fila de datos (desde 1 hasta max_column)
    if status == "green":
        fill = FILL_GREEN
    elif status.startswith("yellow"):
        fill = FILL_YELLOW
    elif status == "red":
        fill = FILL_RED
    else:
        fill = FILL_GRAY
    for col_idx in range(1, ws.max_column + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill

# ---------------------------
# Main
# ---------------------------

def main():
    if not BASE_NORMALIZED.exists():
        print(f"[ERROR] No existe {BASE_NORMALIZED}. Ejecuta primero normalize_names.py", file=sys.stderr)
        sys.exit(1)

    # Cargamos base principal
    base_df = pd.read_excel(BASE_NORMALIZED)

    # Si existe un patched previo, lo usamos para heredar 'status'
    prev_status_map = None
    if BASE_PATCHED_PREV.exists():
        try:
            prev_df = pd.read_excel(BASE_PATCHED_PREV)
            # Clave preferida para mapear: 'arkId' si existe; si no, el índice
            key_col = pick_first_col(prev_df, ["arkId", "arkID", "ark", "id"])
            status_col_prev = pick_first_col(prev_df, ["status", "Status", "STATUS"])
            if key_col and status_col_prev and key_col in base_df.columns:
                prev_status_map = dict(zip(prev_df[key_col].astype(str), prev_df[status_col_prev].astype(str)))
        except Exception as e:
            print(f"[WARN] No se pudo leer {BASE_PATCHED_PREV}: {e}", file=sys.stderr)

    # Identifica columnas clave
    ark_col = pick_first_col(base_df, ["arkId", "arkID", "ark", "id"])
    status_col = pick_first_col(base_df, ["status", "Status", "STATUS"])
    work_name_col = pick_first_col(base_df, ["fullName__work", "fullName_norm", "fullName"])

    # Asegura columna 'status'
    if status_col is None:
        status_col = "status"
        base_df[status_col] = "gray"  # por defecto
    # Hereda status desde patched previo (si se puede)
    if prev_status_map and ark_col:
        # solo rellenamos donde está vacío o por defecto
        def _inherit(row):
            if pd.notna(row[status_col]) and str(row[status_col]).strip().lower() != "gray":
                return row[status_col]
            k = str(row[ark_col]) if pd.notna(row[ark_col]) else None
            if k and k in prev_status_map:
                return prev_status_map[k]
            return row[status_col]
        base_df[status_col] = base_df.apply(_inherit, axis=1)

    # Carga sinónimos y listas
    strong_syns, _weak_syns = load_synonyms(SURNAME_SYNS_CSV) if SURNAME_SYNS_CSV else (set(), set())
    whitelist = load_listfile(WHITELIST_TXT) if WHITELIST_TXT else set()
    blacklist = build_blacklist()

    # Forzado
    promotions = []  # para OUT_FORCE
    n_preserve_green = 0
    n_promoted = 0
    n_ambiguous = 0
    n_blacklisted_excluded = 0

    # Precompute fullname_norm por fila
    fullname_norms = []
    for _, row in base_df.iterrows():
        fullname_norms.append(build_fullname_norm(row, [work_name_col] if work_name_col else []))
    base_df["fullName__norm_tmp"] = fullname_norms  # columna temporal interna

    # Recorre filas y decide promoción
    for idx, row in base_df.iterrows():
        curr_status = str(row[status_col]).strip().lower() if pd.notna(row[status_col]) else "gray"
        fullname_norm = row["fullName__norm_tmp"]

        # whitelists exactas: equivalentes fuertes (pero solo las usamos como extra-seguro)
        if any(re.search(rf"\b{re.escape(w)}\b", fullname_norm) for w in whitelist):
            # NO forzamos automáticamente: solo permite reconocimiento; el forzado lo decide should_force
            pass

        force, reason, ambiguous = should_force(fullname_norm, strong_syns, blacklist)

        if curr_status == "green":
            n_preserve_green += 1
            continue  # nunca degradar

        if ambiguous:
            # marcamos amarillo:ambiguous si no era green
            if not curr_status.startswith("yellow"):
                base_df.at[idx, status_col] = "yellow:ambiguous"
            n_ambiguous += 1
            continue

        if not force:
            # si está claramente en blacklist sin 'zamacona', cuenta como excluido informativo
            if "skip:blacklist" in reason:
                n_blacklisted_excluded += 1
            # no cambiar status
            continue

        # Forzar a green
        base_df.at[idx, status_col] = "green"
        n_promoted += 1

        # Para OUT_FORCE
        force_row = {
            "row_index": idx,
            "arkId": (str(row[ark_col]) if (ark_col and pd.notna(row[ark_col])) else ""),
            "fullName__work": (str(row[work_name_col]) if (work_name_col and pd.notna(row[work_name_col])) else ""),
            "reason": reason
        }
        promotions.append(force_row)

    # Limpieza columna temporal
    base_df.drop(columns=["fullName__norm_tmp"], inplace=True, errors="ignore")

    # Guardar CSV simple
    base_df.to_csv(OUT_CSV, index=False, encoding="utf-8")
    print(f"[OK] {OUT_CSV.name}")

    # Guardar XLSX con colores (si openpyxl disponible)
    base_df.to_excel(OUT_XLSX, index=False)
    print(f"[OK] {OUT_XLSX.name}")

    if load_workbook:
        try:
            wb = load_workbook(OUT_XLSX)
            ws = wb.active
            # encab: fila 1; datos desde fila 2
            # localiza índice de 'status'
            hdr = [c.value for c in ws[1]]
            try:
                status_col_idx = hdr.index(status_col) + 1
            except ValueError:
                status_col_idx = None

            for r in range(2, ws.max_row + 1):
                st = "gray"
                if status_col_idx:
                    cell_v = ws.cell(row=r, column=status_col_idx).value
                    st = str(cell_v).strip().lower() if cell_v is not None else "gray"
                apply_row_fill(ws, r, st if st else "gray")

            wb.save(OUT_XLSX)
        except Exception as e:
            print(f"[WARN] No se pudo aplicar color en XLSX: {e}", file=sys.stderr)

    # Guardar OUT_FORCE (solo filas realmente promovidas ahora)
    with open(OUT_FORCE, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t")
        w.writerow(["row_index", "arkId", "fullName__work", "token", "reason"])
        for r in promotions:
            w.writerow([r["row_index"], r["arkId"], r["fullName__work"], "force:zamacona", r["reason"]])

    print(f"[OK] {OUT_FORCE.name}  ({len(promotions)} filas forzadas a verde)")
    if n_blacklisted_excluded:
        print(f"[INFO] Excluidas {n_blacklisted_excluded} filas por contener blacklist (p.ej. Zamacola/Zamalloa).")
    if n_ambiguous:
        print(f"[INFO] Marcadas {n_ambiguous} filas como yellow:ambiguous (coexistencia con blacklist).")
    print(f"[INFO] Celdas preservadas green: {n_preserve_green} | promovidas ahora: {n_promoted}")

if __name__ == "__main__":
    pd.options.display.width = 200
    pd.options.mode.chained_assignment = None  # silenciar SettingWithCopy
    main()