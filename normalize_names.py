#!/usr/bin/env python3
# Normalización + split robusto + columnas junto al __work + colores + hyperlink en arkId
# Mantiene 100% tu lógica original y añade:
#  - Entrada flexible (prepared.xlsx o all.xlsx)
#  - Columna 'status' derivada de blacklistFlag/reviewFlag (solo informativa)

import re
import unicodedata
from pathlib import Path
import pandas as pd
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

# ---- Entrada flexible: usa prepared si existe; si no, cae a all.xlsx ----
IN_FILE_PREPARED = Path("out/Zamacona_prepared.xlsx")
IN_FILE_ALL      = Path("out/Zamacona_all.xlsx")
if IN_FILE_PREPARED.exists():
    IN_FILE = str(IN_FILE_PREPARED)
elif IN_FILE_ALL.exists():
    IN_FILE = str(IN_FILE_ALL)
else:
    IN_FILE = "out/Zamacona_prepared.xlsx"  # fallback igual que antes

OUT_DIR  = Path("out")
OUT_XLSX = OUT_DIR / "Zamacona_normalized.xlsx"
OUT_CSV  = OUT_DIR / "Zamacona_normalized.csv"
OUT_REVIEW_LOG = OUT_DIR / "Zamacona_review_log.txt"
OUT_UNIQUE_GIVEN = OUT_DIR / "Zamacona_unique_given.txt"
OUT_UNIQUE_SURN  = OUT_DIR / "Zamacona_unique_surnames.txt"

TARGET_PREFIXES = ["fullName","fatherFullName","motherFullName","spouseFullName","childrenFullNames","otherFullNames"]

# ---------- equivalencias de pila ----------
GIVEN_MAP = {
    "joan":"Juan","jhoan":"Juan","jhoanes":"Juan","juo":"Juan","jua":"Juan",
    "josef":"Jose","joseph":"Jose","josephe":"Jose","josepho":"Jose","josefh":"Jose",
    "josepha":"Josefa","jossepha":"Josefa","jpha":"Josefa",
    "ysabel":"Isabel","ysabela":"Isabela","ysabella":"Isabella","isavel":"Isabel",
    "ysavel":"Isabel",
    "ygnacio":"Ignacio","ynacio":"Ignacio","ignazio":"Ignacio","ignazia":"Ignacia","ig":"Ignacio",
    "ysidro":"Isidro","ysidoro":"Isidoro","isidoro":"Isidoro",
    "ypolito":"Hipolito",
    "anttonio":"Antonio","anto":"Antonio","antomio":"Antonio",
    "frco":"Francisco","fr.":"Francisco","francº":"Francisco","franciso":"Francisco","franzisco":"Francisco","frano":"Francisco","franco":"Francisco",
    "cathalina":"Catalina","catarina":"Catalina","chatalina":"Catalina",
    "theresa":"Teresa","theressa":"Teresa","teressa":"Teresa","treresa":"Teresa",
    "michaela":"Micaela",
    "rossa":"Rosa",
    "luzia":"Lucia",
    "ursola":"Ursula",
    "raphael":"Rafael",
    "bartholome":"Bartolome","bartolomeu":"Bartolome",
    "phelipe":"Felipe","philip":"Felipe","philipa":"Felipa","phelipa":"Felipa",
    "matheo":"Mateo","mathia":"Matias","mathea":"Matias","mathias":"Matias",
    "vizente":"Vicente",
    "domenja":"Dominga",
    "mari":"Maria",
    "ma":"Maria",
    "Sebastian":"Sebastian",

    # variantes/typos
    "jagme":"Jaime",
    "crispina":"Crispina","chrispina":"Crispina",
    "ygnacia":"Ignacia","ynes":"Ines",
    "bapptista":"Baptista",
    "asencia":"Ascencia",
    "antonlin":"Antolin",
    "bentura":"Ventura",
    "victora":"Victoria",
    "faustta":"Fausta",
    "marie":"Maria",
    "melatona":"Melitona",
    "asumpcion":"Asuncion",
    "venttura":"Ventura",
    "benrta":"Benita",
    "facunda":"Facundo",
    "ianuario":"Januario",
    "januario":"Januario",
    "yanuario":"Januario",
    "xime":"Ximeno",
    "balentin":"Valentin",
    "presvitero":"Presbitero",
    "ant":"Antonio","antoinio":"Antonio","anttonia":"Antonia",
    "eusevia":"Eusebia",
    "ambrocio":"Ambrosio",
    "manl":"Manuel",
}

# variantes -> Zamacona explícitas (solo estas)
WHITELIST_TOKENS = {"samacona":"Zamacona","camacona":"Zamacona","çamacona":"Zamacona","zamaxa":"Zamacona","zamoana":"Zamacona"}

# NO convertir a Zamacona aunque estén cerca
PROTECTED_NEAR = {"camacana","zamaloba","zamaena","zamagona","samacola","zamacola"}

# ---------- sinónimos de apellidos (minúsculas) -> canónico ----------
SURNAME_SYNONYMS = {
    "birilquicha":"Birisquieta","birizqueta":"Birisquieta",
    "ubirichaga":"Ubiritxaga",
    "herrementeria":"Errementeria",
    "balois":"Valois","baloisoriundo":"Valois",
    "chatave":"Txatabe",
    "sanches":"Sanchez",
    "aierro":"Aiero",
    "basave":"Basabe",
    "heyzaga":"Eizaga","eyzaga":"Eizaga",
    "assua":"Asua",
    "hordenana":"Ordeñana","ordenana":"Ordeñana","ordenana.":"Ordeñana",
    "artuis":"Artiz","jortis":"Ortiz",
    "artamonis":"Artemuniz","artemoniz":"Artemuniz","artamoniz":"Artemuniz",
    "gonzales":"Gonzalez",
    "orve":"Orbe",
    "laragoti":"Larragoiti",
    "artave":"Artabe",
    "echavari":"Etxebarria","echevarria":"Etxebarria","echevarria,":"Etxebarria","echebarria":"Etxebarria",
    "beascoechea":"Beaskoetxea","beacoechea":"Beaskoetxea",
    "vengoechea":"Bengoechea",
    "ozerin":"Ocerin",
    "pagalday":"Pagaldai",
    "vgalde":"Ugalde",
    "zagarduy":"Sagarduy",
    "thelleria":"Telleria",
    "jimeno":"Ximeno",
    "selguera":"Salguera","salguesta":"Salguera",
    "sorzon":"Sorzano",
    "laviador":"Laviada",
    "acion":"Asion",
    "veascoechea":"Beascoechea",
    "velategui":"Belategui",
    "larrizquitu":"Larrizqueta",
    "echartia":"Etxartia",
    "e.":"Eizaga","e":"Eizaga",
    "yrquiza":"Urquiza","irquiza":"Urquiza",
    "aabiagabastia":"Abiagabastida",
    "lasquiban":"Lazquibar",
    "oynquina":"Oiquina",
    "zimpa":"Zimpia",
    "ealo":"Healo",
    "ibierro":"Ibirro",
    "andonay":"Andonaegi",
    "fernandiz":"Fernandez",
    "messo":"Meso",
    "gabon":"Gabona",
    "zoa":"Zua",
    "cenegorta":"Zenegorta",
    "iragori":"Iragorri",
    "urgoytia":"Urgoitia",
    "larrisquitu":"Larruskitu",
    "ochubiaga":"Otxobiaga",
    "eloriaga":"Elorriaga",
    "totorica":"Totorika",
    "perea":"Perez",
    "bequeazuazo":"Bequea",
    "ybertuzca":"Ibertucha",
}

# nombres comunes (para no confundirlos como apellidos)
GIVEN_COMMON = {
    "maria","jose","juan","francisco","manuel","antonio","josefa","francisca","miguel","isabel",
    "teresa","ignacio","martin","pedro","andres","esteban","gabriel","catalina","ana","sebastian",
    "magdalena","manuela","ramon","dominga","rafael","xavier","gregoria","ursula","vicente","agustin"
}

# --- tokens que, si van de SEGUNDO en 2-palabras o al FINAL en >=3, tratamos como SEGUNDO NOMBRE (no apellido)
SECOND_NAME_LIKE = {
    # base comunes
    "isabel","januario","melitona","manuela","francisco","eusebia","ignacia","ambrosio","sebastian",
    # de tus fallos
    "donalo","dano",
}

# blacklist
BLACKLIST_TOKENS = {
    "zamalloa","zamacola","zama","zamboa","zamosa","zamolla","zamolloa","zamallua","zamallao",
    "zamesa","zamezala","zamecilla","zamalla","zamaba","zamoula","zamalea","zamiagua","zamoraya",
    "zamakera","zamcella","zecona","zincacelaya","zalamasa","zepoala","zeamalloa","comasy",
    "zaraguela","jph",
    "camunca","camca","camura","camondas","cambolica","camoca","camuyca",
    "camossa","camdoras","camos","cambamca","camuras","camusca","cameroca",
    "camoresca","cambrenca","cambaca","camocas"
}
BLACKLIST_PHRASES = {"ma na","domingo gar"}
BLACKLIST_REGEX = [
    r"\bcama[a-z]{1,}\b",
    r"\bzamboa\b", r"\bzamosa\b",
    r"\bzamoll?o?a\b",
    r"\bzamall(?:oa|ua|ao)\b",
    r"\bzamecilla\b", r"\bzamalla\b", r"\bzamezala\b", r"\bzamiagua\b",
    r"\bzamesa\b",
    r"\bzam[ou]la\b",
    r"\bzac[a-z]{2,}\b",
    r"\bzay\w+\b",
    r"\bzam(?!acona\b)\w+\b",
]

# colores
FILL_GREEN  = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_YELLOW = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
FILL_GRAY   = PatternFill(fill_type="solid", start_color="E7E6E6", end_color="E7E6E6")

SURNAME_CANON = set()

def strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s or "") if unicodedata.category(c) != "Mn")

def clean_spaces(s: str) -> str:
    s = (s or "").replace("_"," ")
    return re.sub(r"\s+", " ", s).strip()

def archaic_y_to_i(token: str) -> str:
    low = token.lower()
    if low in {"y", "de", "del", "la"}:
        return ""
    return token

def edit_distance(a: str, b: str) -> int:
    a = a.lower(); b = b.lower()
    dp = list(range(len(b)+1))
    for i, ca in enumerate(a, start=1):
        prev = dp[0]; dp[0] = i
        for j, cb in enumerate(b, start=1):
            tmp = dp[j]
            dp[j] = min(dp[j]+1, dp[j-1]+1, prev + (0 if ca==cb else 1))
            prev = tmp
    return dp[-1]

PATTERN_SUBS = [
    (re.compile(r"\bma\s+el\b", re.IGNORECASE), "Maria Elena"),
    (re.compile(r"\bant\.\b", re.IGNORECASE), "Antonio"),
    (re.compile(r"\bmaria\s+antonio\b", re.IGNORECASE), "Maria Antonia"),
    (re.compile(r"\bpresvitero\b", re.IGNORECASE), "Presbitero"),
    (re.compile(r"\b([0-9]+)\b"), ""),
    (re.compile(r"\beme\s+terio\b", re.IGNORECASE), "Emeterio"),
]

def is_surname_like(tok_low: str) -> bool:
    if not tok_low: return False
    if tok_low == "zamacona": return True
    if tok_low in SURNAME_SYNONYMS: return True
    if tok_low in SURNAME_CANON: return True
    if tok_low in GIVEN_COMMON: return False
    return False

def canonicalize_surname(tok: str) -> str:
    if not tok: return ""
    low = strip_accents(tok).lower()
    if low in SURNAME_SYNONYMS:
        return SURNAME_SYNONYMS[low]
    return re.sub(r"\s+", " ", tok.strip()).title()

def dedupe_consecutive(tokens):
    out = []
    for t in tokens:
        if not out or strip_accents(out[-1]).lower() != strip_accents(t).lower():
            out.append(t)
    return out

def merge_compound_given(tokens):
    if not tokens:
        return tokens
    MALE_FIRST = {"francisco","jose","juan","pedro","manuel"}
    low = [strip_accents(t).lower() for t in tokens]
    if len(tokens) >= 2 and low[1] == "antonia" and low[0] in MALE_FIRST:
        tokens = [tokens[0], "Antonio"] + tokens[2:]
        low = [strip_accents(t).lower() for t in tokens]

    COMPOUND_3 = {
        ("francisco","xavier","jesus"),
        ("maria","juana","agustina"),
        ("maria","concepcion","nicasia"),
        ("juana","baptista","geronima"),
        ("mariano","francisco","jaime","sebastian"),
    }
    COMPOUND_2 = {
        ("san","juan"),
        ("maria","antonia"), ("maria","antolin"), ("maria","ascension"), ("maria","asuncion"),
        ("maria","benita"), ("maria","bentura"), ("maria","ventura"), ("maria","dolores"),
        ("maria","elena"), ("maria","fausta"), ("maria","francisca"), ("maria","gregoria"),
        ("maria","isabel"), ("maria","jesus"), ("maria","josefa"), ("maria","manuela"),
        ("maria","melitona"),
        ("juana","melitona"),
        ("maria","victoria"),
        ("jose","antonio"), ("jose","ramon"), ("jose","januario"),
        ("josefa","antonia"),
        ("ana","isabel"),
        ("pedro","antonio"), ("manuel","antonio"),
        ("juan","manuel"), ("francisco","xavier"), ("francisco","ramon"), ("francisco","antonio"),
        ("francisco","juan"),
        ("martin","geronimo"),
        ("francisca","paula"),
        ("felipa","toribia"),
        ("juan","antonio"), ("brigida","dionisia"),
        ("domingo","mariano"), ("miguel","antonio"),
        ("juan","ignacio"), ("juan","jose"),
        ("martin","angel"),
        ("maria","rosa"),
        ("jaime","sebastian"),
        ("alaria","antonia"),
        ("juan","andres"),
        ("luciano","adolfo"),
        ("maria","joanes"),
        ("maria","magdalena"),
        ("juan","domingo"),
        ("juana","josefa"),
        ("juan","anacleto"),
        ("leonor","jose"),
        ("pedro","miguel"),
        ("maria","zeferina"),
        ("cirila","romana"),
        ("cosme","damian"),
        ("manuel","valentin"),
        ("pedro","marcelino"),
        ("juan","domingo"),

        # del log
        ("maria","ascencia"),
        ("esperanza","eusebia"),
        ("dominga","crispina"),
        ("maria","severiana"),
        ("josefa","ignacia"),
        ("juan","ambrosio"),
        ("catalina","francisca"),
        ("eustasia","micaela"),
        ("pedro","ignacio"),
        ("mateo","manuel"),
        ("manuel","jose"),
        ("francisco","sabra"),
        ("manuel","francisco"),
        ("josefa","ibertucha"),
        ("dominga","manuela"),

        # otros reportados
        ("buenaventura","artuisa"),
        ("martin","antonio"),
        ("jose","facundo"),
        ("maria","petra"),
        ("maria","catalina"),
        ("toribia","juliana"),
        ("josefa","francisca"),
        ("josefa","motorn"),
        ("maria","teresa"),
        ("dorotea","alonso"),
        ("vicenta","ines"),
        ("juana","baptista"),
    }

    # 3-palabras al inicio
    if len(tokens) >= 3 and tuple(low[:3]) in COMPOUND_3:
        merged = " ".join([tokens[0].title(), tokens[1].title(), tokens[2].title()])
        rest = tokens[3:]
        if len(rest) == 1 and strip_accents(rest[0]).lower() in SECOND_NAME_LIKE.union(GIVEN_COMMON):
            merged = merged + " " + rest[0].title()
            rest = []
        return [merged] + rest

    # 2-palabras al inicio
    if len(tokens) >= 2 and tuple(low[:2]) in COMPOUND_2:
        merged = " ".join([tokens[0].title(), tokens[1].title()])
        return [merged] + tokens[2:]

    return tokens

TITLE_TOKENS = {"capito","capitan","capitán"}

def normalize_person_item(s: str) -> str:
    if not isinstance(s, str): return ""
    s = strip_accents(s)
    s = re.sub(r"\bcapito\.?\b", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"[.?]+", " ", s)
    s = re.sub(r"[-–—]+", " ", s)
    for rx, repl in PATTERN_SUBS:
        s = rx.sub(repl, s)
    s = clean_spaces(s)
    if not s: return ""
    s = re.sub(r"\bdel\s+", " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bde\s+",  " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\bo\s+",   " ", s, flags=re.IGNORECASE)
    s = re.sub(r"\by\s+",   " ", s, flags=re.IGNORECASE)

    tokens = [t for t in re.split(r"\s+", s) if t]
    norm_tokens = []
    for t in tokens:
        t2 = archaic_y_to_i(t)
        if not t2: continue
        low = t2.lower()
        if low in TITLE_TOKENS: continue
        t3 = GIVEN_MAP.get(low, t2)
        norm_tokens.append(t3)

    norm_tokens = dedupe_consecutive(norm_tokens)

    out = []
    for t in norm_tokens:
        low = strip_accents(t).lower()
        if low in SURNAME_SYNONYMS:
            out.append(SURNAME_SYNONYMS[low]); continue
        if low in WHITELIST_TOKENS:
            out.append(WHITELIST_TOKENS[low]); continue
        if low not in BLACKLIST_TOKENS and low not in PROTECTED_NEAR and edit_distance(low,"zamacona") <= 2:
            out.append("Zamacona"); continue
        out.append(t)
    return clean_spaces(" ".join(out))

def normalize_cell_value(val: str) -> str:
    if not isinstance(val, str): return ""
    parts = [p.strip() for p in val.split(";")]
    normed = [normalize_person_item(p) for p in parts if p]
    return "; ".join([p for p in normed if p])

def contains_token_blacklist(text: str) -> bool:
    toks = re.split(r"[^\wñÑ]+", strip_accents(text or "").lower())
    return any(t in BLACKLIST_TOKENS for t in toks if t)

def contains_regex_blacklist(text: str) -> bool:
    txt = strip_accents(text or "").lower()
    return any(re.search(p, txt) for p in BLACKLIST_REGEX)

def contains_phrase_blacklist(text: str) -> bool:
    txt = clean_spaces(strip_accents(text or "").lower())
    return any(phrase in txt for phrase in BLACKLIST_PHRASES)

def contains_zamacona(text: str) -> bool:
    return "zamacona" in (text or "").lower()

def _tokens(s: str):
    s = (s or "").strip()
    return [t for t in re.split(r"\s+", s) if t] if s else []

def split_person(person: str):
    """Partición robusta con tratamiento de segundos nombres y fallbacks seguros."""
    toks = _tokens(person)
    if not toks:
        return ("", "", "")

    # fusiona compuestos al inicio
    toks = merge_compound_given(toks)
    n = len(toks)
    low = [strip_accents(t).lower() for t in toks]

    # 0) casos triviales
    if n == 1:
        if low[0] == "zamacona":
            return ("", "Zamacona", "")
        return (toks[0].title(), "", "")

    # 1) exactamente 2 tokens
    if n == 2:
        pair = (low[0], low[1])
        if (low[1] in SECOND_NAME_LIKE.union(GIVEN_COMMON)) or pair in {
            ("ana","isabel"), ("jose","januario"), ("juana","melitona"),
            ("dominga","manuela"), ("manuel","francisco"), ("esperanza","eusebia"),
            ("josefa","ignacia"), ("juan","ambrosio"), ("emeterio","donalo")
        }:
            # los 2 tokens son given compuesto → sin apellidos
            return (" ".join([t.title() for t in toks]), "", "")
        # regla estándar 2 → given + primer apellido
        return (toks[0].title(), canonicalize_surname(toks[1]), "")

    # 2) n >= 3 y el ÚLTIMO parece segundo nombre (p.ej. ... Dano/Donalo/Sebastian)
    if low[-1] in SECOND_NAME_LIKE.union(GIVEN_COMMON):
        prev = toks[:-1]           # quitamos el último (lo sumaremos al given)
        last = toks[-1].title()

        if len(prev) == 1:
            # solo queda 1 token antes → todo es given, sin apellidos
            return (f"{prev[0].title()} {last}", "", "")
        if len(prev) == 2:
            p2 = (strip_accents(prev[0]).lower(), strip_accents(prev[1]).lower())
            if p2[1] in SECOND_NAME_LIKE.union(GIVEN_COMMON):
                # también son 2º nombres → todo given
                return (f"{prev[0].title()} {prev[1].title()} {last}", "", "")
            # si no, trata prev[1] como apellido
            return (f"{prev[0].title()} {last}", canonicalize_surname(prev[1]), "")

        # quedan >=3 → seguimos con heurística usando 'prev' y al final añadimos 'last' al given
        toks = prev
        n = len(toks)
        low = [strip_accents(t).lower() for t in toks]
        tail_given = last
    else:
        tail_given = ""  # nada que añadir al given al final

    # 3) heurística estándar n>=3: buscar apellidos desde el final
    idxs = []
    for i in range(n - 1, -1, -1):
        if is_surname_like(low[i]):
            if not idxs or i != idxs[-1]:
                idxs.append(i)
        if len(idxs) == 2:
            break

    # fallback seguro
    if len(idxs) < 2:
        if n >= 2:
            s2 = canonicalize_surname(toks[-1])
            s1 = canonicalize_surname(toks[-2])
            given_core = " ".join(t.title() for t in toks[:-2])
            given = (given_core + (" " + tail_given if tail_given else "")).strip()
            return (given, s1, s2)
        else:
            given = (toks[0].title() + (" " + tail_given if tail_given else "")).strip()
            return (given, "", "")

    s2_idx, s1_idx = idxs[0], idxs[1]
    s2 = canonicalize_surname(toks[s2_idx])
    s1 = canonicalize_surname(toks[s1_idx])
    given_core = " ".join(t.title() for j, t in enumerate(toks) if j < s1_idx)
    given = (given_core + (" " + tail_given if tail_given else "")).strip()
    return (given, s1, s2)

def first_person(cell: str) -> str:
    if not isinstance(cell, str): return ""
    return cell.split(";")[0].strip()

# --- hyperlink & colores ---
def add_ark_hyperlinks(ws):
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    cidx = headers.get("arkId")
    if not cidx: return 0
    letter = get_column_letter(cidx)
    count = 0
    for r in range(2, ws.max_row + 1):
        cell = ws[f"{letter}{r}"]
        val = str(cell.value).strip() if cell.value is not None else ""
        if val.startswith("ark:/61903/1:1:"):
            cell.hyperlink = f"https://www.familysearch.org/{val}?lang=es"
            cell.style = "Hyperlink"
            count += 1
    return count

def paint_rows(ws):
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
    c_black = headers.get("blacklistFlag")
    c_rev   = headers.get("reviewFlag")
    if not c_black or not c_rev:
        return (0,0,0)
    def to_int(v):
        try:
            if v is None: return 0
            if isinstance(v, bool): return int(v)
            return int(str(v).strip())
        except Exception:
            return 0
    green = yellow = gray = 0
    for r in range(2, ws.max_row + 1):
        is_black  = to_int(ws.cell(row=r, column=c_black).value) == 1
        is_review = to_int(ws.cell(row=r, column=c_rev).value) == 1
        fill = FILL_GRAY if is_black else (FILL_YELLOW if is_review else FILL_GREEN)
        if is_black: gray += 1
        elif is_review: yellow += 1
        else: green += 1
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill
    return (green, yellow, gray)

def load_surname_whitelist():
    global SURNAME_CANON
    p = Path("data/whitelist_surnames.txt")
    if not p.exists():
        SURNAME_CANON = set(); return
    vals = []
    for line in p.read_text(encoding="utf-8").splitlines():
        t = strip_accents(line).strip().lower()
        if t: vals.append(t)
    SURNAME_CANON = set(vals)

# ---------- MAIN ----------
def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if not Path(IN_FILE).exists():
        raise SystemExit(f"No encuentro {IN_FILE}")

    load_surname_whitelist()

    df = pd.read_excel(IN_FILE, dtype=str)
    df.columns = [re.sub(r"\s+"," ", str(c)).strip() for c in df.columns]

    work_cols = [c for c in df.columns if c.endswith("__work") and any(c.startswith(p) for p in TARGET_PREFIXES)]
    work_cols.sort()
    if not work_cols:
        raise SystemExit("No se encontraron columnas __work para procesar.")

    # 1) normaliza TODOS los __work
    for c in work_cols:
        df[c] = df[c].fillna("").astype(str).map(normalize_cell_value)

    # 2) flags
    def row_flags_with_reason(row):
        texts = [str(row[c] or "") for c in work_cols]
        has_bad_cam      = any(re.search(r"\bcama[a-z]{1,}\b", strip_accents(t).lower()) for t in texts)
        has_bad_zamregex = any(re.search(r"\bzam(?!acona\b)\w+\b", strip_accents(t).lower()) for t in texts)
        has_bad_regex    = any(contains_regex_blacklist(t) for t in texts)
        has_bad_token    = any(contains_token_blacklist(t) for t in texts)
        has_bad_phrase   = any(contains_phrase_blacklist(t) for t in texts)
        has_z            = any(contains_zamacona(t) for t in texts)

        reason = ""
        if has_bad_token:      reason = "token"
        elif has_bad_cam:      reason = "regex:cama*"
        elif has_bad_zamregex: reason = "regex:zam*!=zamacona"
        elif has_bad_regex:    reason = "regex:misc"
        elif has_bad_phrase:   reason = "phrase"

        critical = has_bad_cam or has_bad_zamregex or has_bad_token or has_bad_phrase
        blacklist = 1 if critical or (has_bad_regex and not has_z) else 0
        review    = 1 if (not blacklist and not has_z) else 0
        return pd.Series({"blacklistFlag": blacklist, "reviewFlag": review, "blacklistReason": reason})

    df = pd.concat([df, df.apply(row_flags_with_reason, axis=1)], axis=1)

    # 2.bis) status informativo (NO afecta tu pintado; sirve a otros scripts)
    if "status" not in df.columns:
        df["status"] = ""
    def _flags_to_status(row):
        b = str(row.get("blacklistFlag","0")).strip()
        r = str(row.get("reviewFlag","0")).strip()
        if b == "1":
            return "gray"
        if r == "1":
            return "yellow"
        return "green"
    df["status"] = df.apply(lambda r: r["status"] if str(r["status"]).strip() else _flags_to_status(r), axis=1)

    # 3) split + columnas
    created = []
    for c in work_cols:
        base = c[:-6]
        series_first = df[c].fillna("").astype(str).map(first_person)
        parts = series_first.map(split_person)
        gcol, s1col, s2col = f"{base}__given", f"{base}__surn1", f"{base}__surn2"
        df[gcol]  = parts.map(lambda t: t[0])
        df[s1col] = parts.map(lambda t: t[1])
        df[s2col] = parts.map(lambda t: t[2])
        created += [gcol, s1col, s2col]

    def reorder():
        order, done = [], set()
        cols = list(df.columns)
        for col in cols:
            if col in done: continue
            order.append(col); done.add(col)
            if col in work_cols:
                base = col[:-6]
                for extra in (f"{base}__given", f"{base}__surn1", f"{base}__surn2"):
                    if extra in df.columns and extra not in done:
                        order.append(extra); done.add(extra)
        return order
    df = df[reorder()]

    # 4) vaciar splits en no-verdes (según tus flags)
    non_green = (df["blacklistFlag"].astype(str)=="1") | (df["reviewFlag"].astype(str)=="1")
    given_cols   = [c for c in created if c.endswith("__given")]
    surname_cols = [c for c in created if c.endswith("__surn1") or c.endswith("__surn2")]
    df.loc[non_green, given_cols] = ""
    df.loc[non_green, surname_cols] = ""

    # 5) guardar
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df.to_csv(OUT_CSV, index=False)
    df.to_excel(OUT_XLSX, index=False)

    # 6) colores + hyperlink arkId (exactamente como ya hacías)
    wb = load_workbook(OUT_XLSX, data_only=False)
    ws = wb.active
    g,y,gr = paint_rows(ws)
    links = add_ark_hyperlinks(ws)
    wb.save(OUT_XLSX)

    # 7) logs (solo verdes por tus flags)
    green = df[(df["blacklistFlag"].astype(str)=="0") & (df["reviewFlag"].astype(str)=="0")].copy()

    name_col = "fullName__work" if "fullName__work" in df.columns else work_cols[0]
    reviews = df[df["reviewFlag"] == 1][name_col].fillna("").astype(str)
    with open(OUT_REVIEW_LOG, "w", encoding="utf-8") as f:
        for v in reviews:
            v = v.strip()
            if v: f.write(v + "\n")

    # únicos (solo verdes)
    def norm_given_string(s: str) -> str:
        s = clean_spaces(s)
        return " ".join(w.capitalize() for w in s.split()) if s else ""

    created_given = [c for c in created if c.endswith("__given")]
    given_strings = []
    for c in created_given:
        given_strings += [norm_given_string(x) for x in green[c].fillna("").astype(str).tolist() if x and x.strip()]
    given_ctr = Counter([g for g in given_strings if g])
    with open(OUT_UNIQUE_GIVEN, "w", encoding="utf-8") as f:
        for tok, cnt in sorted(given_ctr.items(), key=lambda x: (-x[1], x[0].lower())):
            f.write(f"{tok}\t{cnt}\n")

    surn_cols_all = [c for c in created if c.endswith("__surn1") or c.endswith("__surn2")]
    surn_tokens = []
    for c in surn_cols_all:
        surn_tokens += re.split(r"\s*;\s*", ";".join(green[c].fillna("").astype(str).tolist()))
    surn_tokens = [t.strip() for t in surn_tokens if t.strip()]
    surn_ctr = Counter(surn_tokens)
    with open(OUT_UNIQUE_SURN, "w", encoding="utf-8") as f:
        for tok, cnt in sorted(surn_ctr.items(), key=lambda x: (-x[1], x[0].lower())):
            f.write(f"{tok}\t{cnt}\n")

    print(f"[OK] {OUT_XLSX}  ({len(df)} filas)  → filas coloreadas: green={g}, yellow={y}, gray={gr}")
    print(f"[OK] hipervínculos en arkId: {links}")
    print(f"[OK] {OUT_CSV}")
    print(f"[OK] {OUT_REVIEW_LOG} ({reviews.astype(bool).sum()} líneas)")
    print(f"[OK] {OUT_UNIQUE_GIVEN} ({len(given_ctr)} nombres únicos, SOLO verdes)")
    print(f"[OK] {OUT_UNIQUE_SURN} ({len(surn_ctr)} apellidos únicos, SOLO verdes)")

if __name__ == "__main__":
    main()