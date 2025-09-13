#!/usr/bin/env python3
# Marca en ROJO los apellidos rechazados (según out/surnames_reject.tsv)
# Entrada:  out/Zamacona_normalized.xlsx  +  out/surnames_reject.tsv
# Salida :  out/Zamacona_mark_rejects.xlsx
# Extra 1:  out/reject_log.txt        (apellidos rechazados únicos)
# Extra 2:  out/reject_hits.tsv       (log detallado con given por cada hit)

import csv
import unicodedata
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

IN_XLSX      = Path("out/Zamacona_normalized.xlsx")
REJECT_TSV   = Path("out/surnames_reject.tsv")
OUT_XLSX     = Path("out/Zamacona_mark_rejects.xlsx")
OUT_LOG_UNIQ = Path("out/reject_log.txt")
OUT_LOG_HITS = Path("out/reject_hits.tsv")   # NUEVO: log detallado con given

COLOR_WHOLE_ROW = False  # True para pintar fila entera

FILL_RED = PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE")

def strip_accents(s: str) -> str:
    if s is None:
        return ""
    return "".join(c for c in unicodedata.normalize("NFD", str(s)) if unicodedata.category(c) != "Mn")

def load_reject_set(tsv_path: Path) -> set:
    if not tsv_path.exists():
        raise SystemExit(f"No encuentro {tsv_path}. Ejecuta antes audit_surnames.py")
    try:
        df = pd.read_csv(tsv_path, sep="\t", dtype=str).fillna("")
        cols = [c.lower() for c in df.columns]
        if "variant" in cols:
            col = df[df.columns[cols.index("variant")]]
        elif "surname" in cols:
            col = df[df.columns[cols.index("surname")]]
        else:
            col = df.iloc[:, 0]
        return {strip_accents(x).strip().lower() for x in col if str(x).strip()}
    except Exception:
        rejects = set()
        with open(tsv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter="\t")
            header = next(reader, None)
            idx = 0
            if header:
                low = [h.strip().lower() for h in header]
                if "variant" in low:
                    idx = low.index("variant")
                elif "surname" in low:
                    idx = low.index("surname")
                else:
                    idx = 0
            for row in reader:
                if not row:
                    continue
                val = strip_accents(row[idx]).strip().lower()
                if val:
                    rejects.add(val)
        return rejects

def main():
    if not IN_XLSX.exists():
        raise SystemExit(f"No encuentro {IN_XLSX}. Ejecuta antes la normalización.")

    reject_set = load_reject_set(REJECT_TSV)
    if not reject_set:
        print("[AVISO] El conjunto de rechazados está vacío; no se marcará nada.")

    wb = load_workbook(IN_XLSX)
    ws = wb.active

    # Mapa cabeceras -> índice de columna (1-based)
    headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1) if cell.value}

    # Columnas de apellidos y sus índices
    surname_cols = [name for name in headers if name.endswith("__surn1") or name.endswith("__surn2")]
    surname_cols_idx = [headers[name] for name in surname_cols]

    if not surname_cols_idx:
        raise SystemExit("No encuentro columnas de apellidos (__surn1/__surn2) en la hoja.")

    # Para buscar el given, necesitamos la columna base__given
    # Ej: fatherFullName__surn1 -> base = fatherFullName -> buscar fatherFullName__given
    given_cols_map = {}
    for col_name in surname_cols:
        base = col_name.replace("__surn1", "").replace("__surn2", "")
        gname = f"{base}__given"
        given_cols_map[col_name] = headers.get(gname, None)

    marked = 0
    unique_hits = set()
    detailed_rows = []  # Para el TSV detallado

    for r in range(2, ws.max_row + 1):
        row_hits = []
        for col_name, cidx in zip(surname_cols, surname_cols_idx):
            val = ws.cell(row=r, column=cidx).value
            if not val:
                continue
            parts = [p.strip() for p in str(val).split(";") if p.strip()]
            for p in parts:
                key = strip_accents(p).lower()
                if key in reject_set:
                    row_hits.append((col_name, p))

        if row_hits:
            # Obtener el 'given' que corresponde a cada columna base
            for col_name, hit in row_hits:
                base = col_name.replace("__surn1", "").replace("__surn2", "")
                gidx = given_cols_map.get(col_name, None)
                given_val = ws.cell(row=r, column=gidx).value if gidx else ""

                # Agregar a detallado (fila, columna, base, given, hit)
                detailed_rows.append({
                    "row": r,
                    "column": col_name,
                    "base": base,
                    "given": (given_val or "").strip(),
                    "surname_hit": hit
                })

                unique_hits.add(hit)

                if not COLOR_WHOLE_ROW:
                    # Pintar solo la celda de apellido que pegó
                    ws.cell(row=r, column=headers[col_name]).fill = FILL_RED
                    marked += 1

            if COLOR_WHOLE_ROW:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = FILL_RED
                marked += 1

    # Guardar Excel marcado
    wb.save(OUT_XLSX)

    # Guardar log único (solo apellidos)
    with open(OUT_LOG_UNIQ, "w", encoding="utf-8") as f:
        for val in sorted(unique_hits, key=lambda x: x.lower()):
            f.write(val + "\n")

    # Guardar log detallado con given por hit
    # Formato TSV: row, column, base, given, surname_hit
    with open(OUT_LOG_HITS, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t")
        writer.writerow(["row", "column", "base", "given", "surname_hit"])
        for d in sorted(detailed_rows, key=lambda x: (x["row"], x["column"], x["surname_hit"].lower())):
            writer.writerow([d["row"], d["column"], d["base"], d["given"], d["surname_hit"]])

    print(f"[OK] Guardado {OUT_XLSX} (apellidos rechazados marcados en rojo).")
    print(f"[OK] Guardado {OUT_LOG_UNIQ} ({len(unique_hits)} apellidos únicos rechazados).")
    print(f"[OK] Guardado {OUT_LOG_HITS} ({len(detailed_rows)} hits con given).")
    print(f"[INFO] {'Filas' if COLOR_WHOLE_ROW else 'Celdas'} marcadas: {marked}")

if __name__ == "__main__":
    main()